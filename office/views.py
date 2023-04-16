from django.shortcuts import resolve_url, redirect, render, get_object_or_404, HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from openpyxl import load_workbook
from django.views.generic import (
    ListView,
    CreateView,
    UpdateView,
    DeleteView,
    DetailView,
)
from django.utils import timezone
from .models import Post, KidPost, Customer, Material
from django.db.models import Q
from .forms import OfficeForm, OfficeKidForm
import requests
import os
from pathlib import Path
import glob
# import win32com.client

# printer function

import xlrd
import xlwt
from xlutils.copy import copy

import datetime

now = datetime.datetime.now()

BASE_DIR = Path(__file__).resolve().parent.parent


class PostListView(ListView):
    model = Post
    template_name = 'office/index.html'
    context_object_name = 'posts'
    ordering = ['customer_deadline']
    

class WinPostCreateView(CreateView):
    model = Post
    # fields = '__all__'
    form_class = OfficeForm
    template_name = 'office/create.html'
    success_url = reverse_lazy('office-index')
    
    def form_valid(self, form):
        # Excelテンプレートを読み込む
        excel = win32com.client.Dispatch("Excel.Application")
        workbook = excel.Workbooks.Open(os.path.abspath('data/temp_prot.xlsx'))
        worksheet = workbook.Worksheets('Sheet1')

        # 入力内容をExcelファイルに書き込む
        res = self.request.POST
        workbook.Range("C3").Value = res['customer']
        workbook.Range("G3").Value = now.strftime("%m/%d (%A)")
        workbook.Range("J3").Value = res['my_company_deadline']
        workbook.Range("N3").Value = res['customer_deadline']
        workbook.Range("D4").Value = res['memo']
        workbook.Range("T3").Value = res['customer']
        workbook.Range("AF3").Value = res['quantity']
        workbook.Range("T4").Value = res['material']
        workbook.Range("U5").Value = res['memo']
        workbook.Range("AD39").Value = res['customer']
        workbook.Range("").Value = res['']

        # PDFファイルを作成する
        output_path = os.path.abspath('path/to/output.pdf')
        workbook.ExportAsFixedFormat(0, output_path)

        # Excelを終了する
        workbook.Close(False)
        excel.Quit()

        # PDFファイルをダウンロードさせる
        with open(output_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename=output.pdf'
        os.remove(output_path)
        return response
    
        excel = win32com.client.Dispatch("Excel.Application")        #Excelを操作するための設定
        file = excel.Workbooks.Open('Excelファイルの絶対パス')
        excel = win32com.client.Dsipatch("Excel.Application")
        file = excel.Workbooks.Open("エクセルファイルの絶対パス(拡張子は.xlsx)")
        file.WorkSheets(BASE_DIR.joinpath('data/excel/temp_prot.xlsx')).Select()
        file.ActiveSheet.ExportAsFixedFormat(0, "output.pdf")
    
class PostCreateView(CreateView):
    model = Post
    # fields = '__all__'
    form_class = OfficeForm
    template_name = 'office/create.html'
    success_url = reverse_lazy('office-index')
    def get_success_url(self):
        # rb = xlrd.open_workbook((BASE_DIR.joinpath(
        #     'data/excel'), 'temp_prot.xlsx'), formatting_info=True, on_demand=True)
        return resolve_url('office-index')
""""
    def form_valid(self, form):
        # self.request.session['POST'] = self.request.POST
        res = self.request.POST

        # Excelファイルにデータを入力する
        wb = xlwt.Workbook()
        print("fuck!!!!!!!")
        rb = load_workbook((BASE_DIR.joinpath('data/excel/temp_prot.xlsx')))
        ws = rb.active
        ws['C3'] = res['customer']
        ws['D4'] = res['memo']
        # PDFファイルを作成する
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="output.pdf"'
        p = canvas.Canvas(response)

        # Excelファイルの内容をPDFに書き込む
        x_offset = 50
        y_offset = 720
        row_height = 20
        for row in ws.iter_rows():
            for cell in row:
                p.drawString(x_offset, y_offset, str(cell.value))
                x_offset += 100
            y_offset -= row_height
            x_offset = 50

        # PDFファイルを保存する
        p.showPage()
        p.save()
        return response
"""
    # Excelのテンプレートファイルの読み込み

    # 読み込んだExcelファイルをコピーする
    # wb = copy(rb)
    # # フォント設定（Excelファイル内部でよく使うものはあらかじめ設定しておいた方が良いです。）
    # font_normal = xlwt.Font()
    # font_normal.name = 'ＭＳ Ｐ明朝'
    # # テキストを書き込む
    # wb.write(row, col, value, style_text)
    # # 日付を書き込む
    # wb.write(row, col, value, style_date)
    # # 通貨を書き込む
    # wb.write(row, col, value, style_currency)
    # # 数値を書き込む
    # wb.write(row, col, value, style_num)
    # # Decimalを書き込む
    # wb.write(row, col, value, style_decimal)

    # # 表示位置設定（左寄せ上下中央揃え）
    # align_normal = xlwt.Alignment()
    # align_normal.horz = xlwt.Alignment.HORZ_LEFT
    # align_normal.vert = xlwt.Alignment.VERT_CENTER

    # # 表示位置設定（右寄せ上下中央揃え）
    # align_right = xlwt.Alignment()
    # align_right.horz = xlwt.Alignment.HORZ_RIGHT
    # align_right.vert = xlwt.Alignment.VERT_CENTER

    # # 罫線設定（全てに細い罫線を引く）
    # border_all = xlwt.Borders()
    # border_all.top = xlwt.Borders.THIN
    # border_all.bottom = xlwt.Borders.THIN
    # border_all.left = xlwt.Borders.THIN
    # border_all.right = xlwt.Borders.THIN

    # # 背景色設定（黄色）
    # pattern = xlwt.Pattern()
    # pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    # # 色と紐付くコードの情報はStyle.pyに記載されています。
    # pattern.pattern_fore_colour = 0x0D

    # # ノーマルなテキストのスタイル
    # style_text = xlwt.XFStyle()
    # style_text.font = font_normal
    # style_text.borders = border_all
    # style_text.alignment = align_normal

    # # ノーマルな日付のスタイル
    # style_date = xlwt.easyxf('font: name ＭＳ Ｐ明朝', 'YYYY年M月D日')
    # style_date.borders = border_all
    # style_date.alignment = align_normal

    # # ノーマルな通貨のスタイル
    # style_currency = xlwt.easyxf('font: name ＭＳ Ｐ明朝', '¥#,##0')
    # style_currency.borders = border_all
    # style_currency.alignment = align_normal

    # # 右寄せの通貨のスタイル
    # style_currency_align_right = xlwt.easyxf('font: name ＭＳ Ｐ明朝', '¥#,##0')
    # style_currency_align_right.borders = border_all
    # style_currency_align_right.alignment = align_right

    # # ノーマルな数値のスタイル
    # style_num = xlwt.easyxf('font: name ＭＳ Ｐ明朝', '#,##0')
    # style_num.borders = border_all
    # style_num.alignment = align_right

    # # ノーマルなDecimalのスタイル
    # style_decimal = xlwt.easyxf('font: name ＭＳ Ｐ明朝', '#,##0.0')
    # style_decimal.borders = border_all
    # style_decimal.alignment = align_right

    # # WorkBookからWorkSheetを取得する
    # ws = wb.get_sheet(sheet_name)



class PostUpdateView(UpdateView):
    model = Post
    form_class = OfficeForm
    template_name = 'office/update.html'

    def get_success_url(self):
        return reverse('office-index')

    def get_form(self):
        form = super(PostUpdateView, self).get_form()
        form.fields['customer'].label = '取引先'

        return form


class PostDetailView(DetailView):
    model = Post
    template_name = 'office/detail.html'


class PostDeleteView(DeleteView):
    model = Post
    template_name = 'office/delete.html'

    def get_success_url(self):
        return reverse('office-index')


""" Office-Kid """
""" PostKidとKidPostの順番を間違えた """


class PostKidListView(ListView):
    model = KidPost
    template_name = 'office-kid/index.html'
    context_object_name = 'post_kids'
    ordering = ['customer_deadline']

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['posts'] = Post.objects.all
        return context


class PostKidCreateView(CreateView):
    model = KidPost
    form_class = OfficeKidForm
    template_name = 'office-kid/create.html'

    def get_success_url(self):
        return resolve_url('office-kid-index')


class PostKidUpdateView(UpdateView):
    model = KidPost
    form_class = OfficeKidForm
    template_name = 'office-kid/update.html'

    def get_success_url(self):
        return reverse('office-kid-index')

    def get_form(self):
        form = super(PostKidUpdateView, self).get_form()
        form.fields['customer'].label = '取引先'
        return form


class PostKidDeleteView(DeleteView):
    model = KidPost
    template_name = 'office-kid/delete.html'

    def get_success_url(self):
        return reverse('office-kid-index')


""" Customer """


class CustomerListView(ListView):
    model = Customer
    template_name = 'customer/index.html'
    context_object_name = 'customers'
    ordering = ['name']


class CustomerCreateView(CreateView):
    model = Customer
    template_name = 'customer/create.html'
    fields = ("name", "email", 'fax', "tel_number_1", "tel_number_2")

    def get_success_url(self):
        return reverse('customer-index')


class CustomerUpdateView(UpdateView):
    model = Customer
    fields = ("name", "email", 'fax', "tel_number_1", "tel_number_2")
    template_name = 'customer/update.html'

    def get_success_url(self):
        return reverse('customer-index')

    def get_form(self):
        form = super(CustomerUpdateView, self).get_form()
        form.fields['name'].label = '取引先'
        return form


class CustomerDeleteView(DeleteView):
    model = Customer
    template_name = 'customer/delete.html'

    def get_success_url(self):
        return reverse('customer-index')


class MaterialListView(ListView):
    model = Material
    template_name = 'material/index.html'
    context_object_name = 'materials'
    ordering = ['name']


class MaterialCreateView(CreateView):
    model = Material
    template_name = 'material/create.html'
    fields = ("name")

    def get_success_url(self):
        return reverse('material-index')


class MaterialUpdateView(UpdateView):
    model = Material
    fields = ("name")
    template_name = 'material/update.html'

    def get_success_url(self):
        return reverse('material-index')

    def get_form(self):
        form = super(MaterialUpdateView, self).get_form()
        form.fields['name'].label = '取引先'
        return form


class MaterialDeleteView(DeleteView):
    model = Material
    template_name = 'material/delete.html'

    def get_success_url(self):
        return reverse('material-index')
